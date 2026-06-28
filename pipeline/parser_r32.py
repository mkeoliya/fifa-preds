"""Parse RO32 re-draft workbooks into Prediction objects (knockout only).

Same Excel layout as the OG template; sheet name is 'World Cup Repreds'.
Only knockout stages (R32 through FINAL) are read — group matches are skipped.
Empty/future slots are silently skipped rather than raising errors.
Outputs to data/picks_r32/.
"""
from __future__ import annotations

import pickle
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

from models import MatchPick, Prediction

RAW_R32_DIR = Path(__file__).resolve().parent.parent / "data" / "Repreds"
PICKS_R32_DIR = Path(__file__).resolve().parent.parent / "data" / "picks_r32"
SHEET_NAME = "World Cup Repreds"

NAME_MAP = {
    "RO32Repreds_Cunt":         "Cunt",
    "RO32Repreds_Frenchie":     "Frenchie",
    "RO32Repreds_JJ":           "JJ",
    "RO32Repreds_Kaku":         "Kaku",
    "RO32Repreds_Kalo":         "Kaalo",
    "RO32Repreds_Kutu":         "Kutu",
    "RO32Repreds_Majank":       "Majank",
    "RO32Repreds_Manna":        "Manna",
    "RO32Repreds_Mayukh":       "Mayukh",
    "RO32Repreds_Sexy Rijo":    "Sexy Rijo",
    "RO32Repreds_Shaun":        "Shaun",
    "RO32Repreds_Snacc":        "Snacc",
    "RO32Repreds_Upa":          "Upa",
    "RO32Repreds_Vui":          "Vui",
}

# Same KO layout as the OG template.
KO_LAYOUT = [
    ("R32",   48, [1 + 3 * i for i in range(16)]),
    ("R16",   58, [1 + 3 * i for i in range(8)]),
    ("QF",    68, [2, 8, 14, 20]),
    ("SF",    78, [5, 17]),
    ("THIRD", 86, [11]),
    ("FINAL", 96, [11]),
]
WIDE_TEAM_GAP_STAGES = {"QF", "SF", "THIRD", "FINAL"}


def _int(v):
    if v is None or v == "":
        return None
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return None


def _date(v):
    return v.strftime("%Y-%m-%d %H:%M") if isinstance(v, datetime) else None


def _cell(ws, row, col):
    return ws.cell(row=row, column=col).value


def _stem(path: Path) -> str:
    """Return the name-map key, stripping trailing _YYYYMMDD suffixes."""
    stem = path.stem
    # strip date suffix like _20260628
    import re
    stem = re.sub(r'_\d{8}$', '', stem)
    return stem


def parse_workbook(path: Path) -> Prediction:
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"{path.name}: sheet '{SHEET_NAME}' not found "
                         f"(available: {wb.sheetnames})")
    ws = wb[SHEET_NAME]
    name = NAME_MAP.get(_stem(path), _stem(path))
    pred = Prediction(name=name, source_file=path.name)

    for stage, row, cols in KO_LAYOUT:
        gap = 2 if stage in WIDE_TEAM_GAP_STAGES else 1
        for col in cols:
            match_no = _int(_cell(ws, row, col))
            t1c, t2c = col + 1, col + 1 + gap
            t1 = _cell(ws, row + 2, t1c)
            t2 = _cell(ws, row + 2, t2c)
            if match_no is None or not t1 or not t2:
                continue  # future round slot not yet filled
            pred.matches.append(MatchPick(
                match_no=match_no, stage=stage,
                team1=str(t1), team2=str(t2),
                score1=_int(_cell(ws, row + 3, t1c)),
                score2=_int(_cell(ws, row + 3, t2c)),
                pen1=_int(_cell(ws, row + 5, t1c)),
                pen2=_int(_cell(ws, row + 5, t2c)),
                date=_date(_cell(ws, row + 1, col + 1)),
                venue=_cell(ws, row, col + 1),
            ))
    return pred


def main() -> int:
    PICKS_R32_DIR.mkdir(parents=True, exist_ok=True)
    failures = []
    parsed = 0
    for path in sorted(RAW_R32_DIR.glob("*.xlsx")):
        stem = _stem(path)
        name = NAME_MAP.get(stem, stem)
        try:
            pred = parse_workbook(path)
        except Exception as e:
            failures.append(f"{path.name}: {e}")
            continue

        r32_count = sum(1 for m in pred.matches if m.stage == "R32")
        print(f"[OK] {name:12s} {r32_count} R32 picks, "
              f"{len(pred.matches)} total KO picks")

        out = PICKS_R32_DIR / f"{name.replace(' ', '_')}.pkl"
        with open(out, "wb") as f:
            pickle.dump(pred, f)
        with open(out, "rb") as f:
            loaded = pickle.load(f)
        assert loaded.to_dict() == pred.to_dict(), f"round-trip mismatch: {name}"
        parsed += 1

    for fmsg in failures:
        print(f"[FAIL] {fmsg}")
    print(f"\nParsed {parsed} re-draft files -> {PICKS_R32_DIR}")
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
