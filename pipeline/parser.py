"""Parse Hermann Baum WC2026 predictor workbooks into Prediction objects.

The template stores everything at fixed coordinates in the 'World Cup' sheet:
  * 72 group matches in a 12-wide x 6-tall grid of 3-column blocks
  * knockout rounds at fixed anchor rows (R32=48, R16=58, QF/SF/3rd/Final below)
Each block: anchor row = (match_no, venue), +1 date, +2 teams, +3 scores,
penalty shootout scores on the row two below the score row (knockouts only).
"""
from __future__ import annotations

import pickle
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from models import AwardPicks, MatchPick, Prediction

RAW_DIR = Path(__file__).resolve().parent.parent / "data" / "raw"
PICKS_DIR = Path(__file__).resolve().parent.parent / "data" / "picks"

# filename stem -> display name
NAME_MAP = {
    "CuntPreds": "Cunt",
    "Frenchie Preds": "Frenchie",
    "JJ Preds": "JJ",
    "KaaloPreds": "Kaalo",
    "KakuPred": "Kaku",
    "KutuPreds": "Kutu",
    "MajankPreds": "Majank",
    "Mayukh_preds": "Mayukh",
    "Sexy Rijo Preds": "Sexy Rijo",
    "Shaun Preds": "Shaun",
    "SnaccPreds": "Snacc",
    "Upa Preds": "Upa",
    "VuiPreds": "Vui",
}

# Award picks supplied outside the workbook (file missing the Awards sheet).
AWARD_OVERRIDES = {
    "Majank": AwardPicks(
        golden_ball="Pedri",
        golden_boot="Erling Haaland",
        golden_glove="Jordan Pickford",
        young_player="Lamine Yamal",
    ),
}
# Per group decision: anyone else with no award picks defaults to Kutu's.
# (Applies to Snacc, Vui — missing Awards sheet — and JJ, Mayukh — blank sheet.)

# Group stage grid: 12 column-blocks, first col of each block holds match_no,
# +1 venue/team1, +2 team2. Row anchors for the 6 match rows per column.
GROUP_BLOCK_COLS = [1 + 3 * i for i in range(12)]  # A, D, G, ... AH
GROUP_ROW_ANCHORS = [11, 16, 21, 26, 31, 36]

# Knockouts: (stage, anchor_row, [first col of each block]).
# Block layout: anchor=(match_no @ col, venue @ col+1), teams @ +2 in
# col+1/col+2, scores @ +3, pens @ +5.
KO_LAYOUT = [
    ("R32", 48, [1 + 3 * i for i in range(16)]),       # A..AT
    ("R16", 58, [1 + 3 * i for i in range(8)]),        # A..V
    ("QF", 68, [2, 8, 14, 20]),                        # B, H, N, T
    ("SF", 78, [5, 17]),                               # E, Q
    ("THIRD", 86, [11]),                               # K
    ("FINAL", 96, [11]),                               # K
]
# In QF/SF/3rd/Final blocks the two team cells are 2 apart, not adjacent.
WIDE_TEAM_GAP_STAGES = {"QF", "SF", "THIRD", "FINAL"}

CHAMPION_CELL = "R98"


def _int(v):
    if v is None or v == "":
        return None
    return int(round(float(v)))


def _date(v):
    return v.strftime("%Y-%m-%d %H:%M") if isinstance(v, datetime) else None


def _cell(ws, row, col):
    return ws.cell(row=row, column=col).value


def parse_workbook(path: Path) -> Prediction:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["World Cup"]
    name = NAME_MAP.get(path.stem, path.stem)
    pred = Prediction(name=name, source_file=path.name)

    # --- group stage ---
    for col in GROUP_BLOCK_COLS:
        for row in GROUP_ROW_ANCHORS:
            match_no = _int(_cell(ws, row, col))
            t1 = _cell(ws, row + 2, col + 1)
            t2 = _cell(ws, row + 2, col + 2)
            if match_no is None or not t1 or not t2:
                raise ValueError(
                    f"{path.name}: empty group match at "
                    f"{get_column_letter(col)}{row}"
                )
            pred.matches.append(MatchPick(
                match_no=match_no, stage="GROUP", team1=str(t1), team2=str(t2),
                score1=_int(_cell(ws, row + 3, col + 1)),
                score2=_int(_cell(ws, row + 3, col + 2)),
                date=_date(_cell(ws, row + 1, col + 1)),
                venue=_cell(ws, row + 1, col + 1) and _cell(ws, row, col + 1),
            ))

    # --- knockout stages ---
    for stage, row, cols in KO_LAYOUT:
        gap = 2 if stage in WIDE_TEAM_GAP_STAGES else 1
        for col in cols:
            match_no = _int(_cell(ws, row, col))
            t1c, t2c = col + 1, col + 1 + gap
            t1 = _cell(ws, row + 2, t1c)
            t2 = _cell(ws, row + 2, t2c)
            if match_no is None or not t1 or not t2:
                raise ValueError(
                    f"{path.name}: empty {stage} match at "
                    f"{get_column_letter(col)}{row}"
                )
            pred.matches.append(MatchPick(
                match_no=match_no, stage=stage, team1=str(t1), team2=str(t2),
                score1=_int(_cell(ws, row + 3, t1c)),
                score2=_int(_cell(ws, row + 3, t2c)),
                pen1=_int(_cell(ws, row + 5, t1c)),
                pen2=_int(_cell(ws, row + 5, t2c)),
                date=_date(_cell(ws, row + 1, col + 1)),
                venue=_cell(ws, row, col + 1),
            ))

    pred.champion = ws[CHAMPION_CELL].value

    # --- awards ---
    if "Awards" in wb.sheetnames:
        aw = wb["Awards"]
        vals = {}
        for r in range(2, 10):
            cat, player = aw.cell(row=r, column=1).value, aw.cell(row=r, column=2).value
            if cat and player:
                vals[str(cat).lower()] = str(player).strip()
        for key, attr in [("ball", "golden_ball"), ("boot", "golden_boot"),
                          ("glove", "golden_glove"), ("young", "young_player")]:
            for cat, player in vals.items():
                if key in cat:
                    setattr(pred.awards, attr, player)
    return pred


def validate(pred: Prediction) -> list[str]:
    problems = []
    counts = {}
    for m in pred.matches:
        counts[m.stage] = counts.get(m.stage, 0) + 1
        if m.score1 is None or m.score2 is None:
            problems.append(f"match {m.match_no} ({m.stage}) missing score")
        if m.stage != "GROUP" and m.score1 == m.score2 and m.pen1 is None:
            problems.append(f"match {m.match_no} ({m.stage}) drawn but no pens")
    expected = {"GROUP": 72, "R32": 16, "R16": 8, "QF": 4, "SF": 2,
                "THIRD": 1, "FINAL": 1}
    for stage, n in expected.items():
        if counts.get(stage) != n:
            problems.append(f"{stage}: {counts.get(stage)} matches, expected {n}")
    nos = sorted(m.match_no for m in pred.matches)
    if nos != list(range(1, 105)):
        problems.append("match numbers are not exactly 1..104")
    if not pred.champion:
        problems.append("no champion")
    if not pred.awards.golden_ball:
        problems.append("no award picks")
    return problems


def main() -> int:
    PICKS_DIR.mkdir(parents=True, exist_ok=True)
    preds: dict[str, Prediction] = {}
    failures = []
    for path in sorted(RAW_DIR.glob("*.xlsx")):
        try:
            preds[NAME_MAP.get(path.stem, path.stem)] = parse_workbook(path)
        except Exception as e:
            failures.append(f"{path.name}: {e}")

    kutu = preds.get("Kutu")
    for name, pred in preds.items():
        if pred.awards.golden_ball is None:
            if name in AWARD_OVERRIDES:
                pred.awards = AWARD_OVERRIDES[name]
            elif kutu:
                pred.awards = kutu.awards

    for name, pred in preds.items():
        problems = validate(pred)
        status = "OK " if not problems else "WARN"
        print(f"[{status}] {name:10s} champion={pred.champion!r} "
              f"golden_ball={pred.awards.golden_ball!r}")
        for p in problems:
            print(f"        - {p}")
        out = PICKS_DIR / f"{name.replace(' ', '_')}.pkl"
        with open(out, "wb") as f:
            pickle.dump(pred, f)
        # round-trip check
        with open(out, "rb") as f:
            loaded = pickle.load(f)
        assert loaded.to_dict() == pred.to_dict(), f"round-trip mismatch: {name}"

    for fmsg in failures:
        print(f"[FAIL] {fmsg}")
    print(f"\nParsed {len(preds)} files -> {PICKS_DIR}")
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
